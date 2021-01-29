import xlrd
import MySQLdb
import os
import openpyxl

def insert():
    book = xlrd.open_workbook(filename = 'C:/Users/Public/Documents/somexeb.xlsx')
    sheet = book.sheet_by_name("Sheet1")

    database = MySQLdb.connect(host="210.179.174.151",user="jwbig",password="EncglsBig!!",db="enc_bigdata",charset="utf8")
    cursor = database.cursor()

    query = """insert into skx_exw2 values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

    print("STart")

    for r in range(1, sheet.nrows):
        UPDUSERCD1 = sheet.cell(r,0).value
        SOKEY = sheet.cell(r,1).value
        DELIVERYCD = sheet.cell(r,2).value
        ACTFLG = sheet.cell(r,3).value
        PICKBATCHKEY = sheet.cell(r,4).value
        ASSYKEY = sheet.cell(r,5).value
        SOTYPE = sheet.cell(r,6).value
        STS = sheet.cell(r,7).value
        ORDERDATE = sheet.cell(r,8).value
        DELIVNAME = sheet.cell(r,9).value
        SHIPSCHDATE = sheet.cell(r,10).value
        SHIPDATE = sheet.cell(r,11).value
        PRIORITIES = sheet.cell(r,12).value
        OWNERCD = sheet.cell(r,13).value
        ITEMCD = sheet.cell(r,14).value
        EXPECTQTY1 = sheet.cell(r,15).value
        SHIPPEDQTY1 = sheet.cell(r,16).value
        CUSTOMERCD = sheet.cell(r,17).value
        OWNERNM = sheet.cell(r,18).value
        ORDERTYPE = sheet.cell(r,19).value
        DELIVSCHDATE = sheet.cell(r,20).value
        DELIVDATE = sheet.cell(r,21).value
        TRANSPORTPRIORITY = sheet.cell(r,22).value
        WAREHOUSECD = sheet.cell(r,23).value
        OTHERREFNO3 = sheet.cell(r,24).value
        WAREHOUSENM = sheet.cell(r,25).value
        PREKEY = sheet.cell(r,26).value
        CUSTOMERNM = sheet.cell(r,27).value
        SHIPTOCD = sheet.cell(r,28).value
        SHIPTONM = sheet.cell(r,29).value
        ADDRESS1 = sheet.cell(r,30).value
        ADDRESS2 = sheet.cell(r,31).value
        POSTNO1 = sheet.cell(r,32).value
        POSTNO2 = sheet.cell(r,33).value
        HPNO = sheet.cell(r,34).value
        EMAIL = sheet.cell(r,35).value
        REFNAME = sheet.cell(r,36).value
        ALLOCGROUP = sheet.cell(r,37).value
        NOTES = sheet.cell(r,38).value
        OTHERREFNO1 = sheet.cell(r,39).value
        OTHERREFNO2 = sheet.cell(r,40).value
        MOVEKEY = sheet.cell(r,41).value
        WBKEY = sheet.cell(r,42).value
        ADDDATETIME = sheet.cell(r,43).value
        ADDUSERCD = sheet.cell(r,44).value
        UPDDATETIME = sheet.cell(r,45).value
        UPDUSERCD = sheet.cell(r,46).value
        TERMINALCD = sheet.cell(r,47).value
        ORDERCUSTNAME = sheet.cell(r,48).value
        ORDERCUSTPHONE = sheet.cell(r,49).value
        ORDERCUSTHP = sheet.cell(r,50).value
        PHONENO = sheet.cell(r,51).value
        SOKEY1 = sheet.cell(r,52).value
        SOLINENO = sheet.cell(r,53).value
        ORDERKEY = sheet.cell(r,54).value
        ORDERLINENO = sheet.cell(r,55).value
        ITEMGROUP = sheet.cell(r,56).value
        IFITEMCD = sheet.cell(r,57).value
        LOT1 = sheet.cell(r,58).value
        LOT2 = sheet.cell(r,59).value
        LOT3 = sheet.cell(r,60).value
        LOT4 = sheet.cell(r,61).value
        LOT5 = sheet.cell(r,62).value
        NOSHIPPINGFLG = sheet.cell(r,63).value
        OTHERFLG = sheet.cell(r,64).value
        XDOCKQTY1 = sheet.cell(r,65).value
        ALLOCQTY1 = sheet.cell(r,66).value
        PICKEDQTY1 = sheet.cell(r,67).value
        SORTEDQTY1 = sheet.cell(r,68).value
        ADJUSTQTY1 = sheet.cell(r,69).value
        PRICE1 = sheet.cell(r,70).value
        PRICE2 = sheet.cell(r,71).value
        PRICE3 = sheet.cell(r,72).value
        NOTES1 = sheet.cell(r,73).value
        LOTRESERVEFLG = sheet.cell(r,74).value
        PICKQTYERRORFLG = sheet.cell(r,75).value 
        XDOCFLG = sheet.cell(r,76).value
        MOVELINENO = sheet.cell(r,77).value
        ASSYLINENO = sheet.cell(r,78).value
        ADDDATETIME1 = sheet.cell(r,79).value
        ADDUSERCD1 = sheet.cell(r,80).value
        UPDDATETIME1 = sheet.cell(r,81).value
        TERMINALCD1 = sheet.cell(r,82).value
        SHOPITEMCD = sheet.cell(r,83).value
        OTHERVALUE1 = sheet.cell(r,84).value
        OTHERVALUE2 = sheet.cell(r,85).value
        FAX1 = sheet.cell(r,86).value
        CONFIRMQTY = sheet.cell(r,87).value
        NAME = sheet.cell(r,88).value
        CONDATE = sheet.cell(r,89).value
        CRUD = sheet.cell(r,90).value
        SODAY = sheet.cell(r,91).value
        IFDATE = sheet.cell(r,92).value
        IFSODAY = sheet.cell(r,93).value

        values = (UPDUSERCD1
                ,SOKEY
                , DELIVERYCD
                , ACTFLG
                , PICKBATCHKEY
                , ASSYKEY
                , SOTYPE
                , STS
                , ORDERDATE
                , DELIVNAME
                , SHIPSCHDATE
                , SHIPDATE
                , PRIORITIES
                , OWNERCD
                , ITEMCD
                , EXPECTQTY1
                , SHIPPEDQTY1
                , CUSTOMERCD
                , OWNERNM
                , ORDERTYPE
                , DELIVSCHDATE
                , DELIVDATE
                , TRANSPORTPRIORITY
                , WAREHOUSECD
                , OTHERREFNO3
                , WAREHOUSENM
                , PREKEY
                , CUSTOMERNM
                , SHIPTOCD
                , SHIPTONM
                , ADDRESS1
                , ADDRESS2
                , POSTNO1
                , POSTNO2
                , HPNO
                , EMAIL
                , REFNAME
                , ALLOCGROUP
                , NOTES
                , OTHERREFNO1
                , OTHERREFNO2
                , MOVEKEY
                , WBKEY
                , ADDDATETIME
                , ADDUSERCD
                , UPDDATETIME
                , UPDUSERCD
                , TERMINALCD
                , ORDERCUSTNAME
                , ORDERCUSTPHONE
                , ORDERCUSTHP
                , PHONENO
                , SOKEY1
                , SOLINENO
                , ORDERKEY
                , ORDERLINENO
                , ITEMGROUP
                , IFITEMCD
                , LOT1
                , LOT2
                , LOT3
                , LOT4
                , LOT5
                , NOSHIPPINGFLG
                , OTHERFLG
                , XDOCKQTY1
                , ALLOCQTY1
                , PICKEDQTY1
                , SORTEDQTY1
                , ADJUSTQTY1
                , PRICE1
                , PRICE2
                , PRICE3
                , NOTES1
                , LOTRESERVEFLG
                , PICKQTYERRORFLG
                , XDOCFLG
                , MOVELINENO
                , ASSYLINENO
                , ADDDATETIME1
                , ADDUSERCD1
                , UPDDATETIME1
                , TERMINALCD1
                , SHOPITEMCD
                , OTHERVALUE1
                , OTHERVALUE2
                , FAX1
                , CONFIRMQTY
                , NAME
                , CONDATE
                , CRUD
                , SODAY
                , IFDATE
                , IFSODAY)
        
        cursor.execute(query,values)
        print(r)
    cursor.close()
    database.commit()
    database.close()
    print ("")
    print ("")
    columns = str(sheet.ncols)
    rows = str(sheet.nrows)
    #print ("Imported", columns,"columns and", rows, "rows. All Done!")
insert()